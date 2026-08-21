"""
Ausgabe im Layout der BSI-Vorlage A.3.4 (Kompendium-Edition 2023).

Zweck: direkte Vergleichbarkeit mit den bestehenden Grundschutz-2023-Vorlagen.
Spaltenreihenfolge, Vokabular, Kopfzeilenposition und Rollup-Semantik sind aus
'A_3_4_Vorlage_Modellierung_IT-Grundschutz-Check_1_1_1.xlsm' uebernommen.

Strukturabbildung GS++ -> 2023
------------------------------
2023 kennt zwei Ebenen: Anforderung (z.B. APP.3.3.A2) und Teilanforderung
(APP.3.3.A2.1..A2.4). Die Anforderungszeile hat kein Eingabefeld; ihr Status
wird aus den Teilanforderungen aggregiert. Genau dieselbe Mechanik hat GS++,
nur mit bis zu vier Ebenen. Abbildung:

  Anforderungszeile   = jeder GS++-Knoten auf Ebene 0        (651 Zeilen)
  Teilanforderungen   = alle pruefpflichtigen Knoten darunter (973 Zeilen)
                        einschliesslich des Ebene-0-Knotens selbst, sofern
                        dieser pruefpflichtig ist

Sammelanforderungen (effort_level 0) tauchen NICHT als Teilanforderung auf -
sie sind reine Klammern; ihre Kinder haengen direkt an der Anforderungszeile.
Damit bleibt die Zahl der bewertbaren Zeilen exakt bei 973.

Ein Blatt je Praktik, analog zu einem Blatt je Baustein.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .excel import (
    EINGABE_FILL,
    FONT,
    KOPF_FILL,
    KOPF_FONT,
    RAHMEN,
    RELEVANZ,
    ROLLUP_FILL,
    UMSETZUNGSSTATUS,
    ZELL_FONT,
    _blatt_metadaten,
    _rollup_formel,
)
from . import design
from .models import CatalogSnapshot, Requirement

log = logging.getLogger(__name__)

KOPFZEILE = 6  # wie in der BSI-Vorlage

SPALTEN_2023 = [
    ("Anforderung", 16),
    ("Beschreibung", 40),
    ("Teilanforderung", 18),
    ("Schutz", 30),
    ("Anforderungstext", 78),
    ("Umsetzung", 30),
    ("Nachweis", 26),
    ("Status", 26),
    ("Maßnahme", 30),
    ("Befragte Person", 18),
    ("Prüfende Person", 18),
    ("Datum der Prüfung", 16),
    # --- Zusatzspalten (nicht Teil der BSI-Vorlage A.3.4, ergaenzt fuer msg) ---
    ("Schwerpunkt (Schutzziel)", 30),
    ("Alte Anforderung (IT-GS 2023)", 40),
    ("Beziehung (IT-GS 2023)", 16),
    # --- Guidance-Block, standardmaessig eingeklappt ---
    ("Erläuterung", 80),
    ("Zielobjekt-Kategorien", 26),
    ("Dokumentation", 26),
    ("Ergebnis", 40),
    ("Gefährdungen", 46),
    ("Aufwand", 9),
    # Schutzziele einzeln: fuer Filter und Pivot, wo die Textspalte nicht traegt
    ("Vertraulichkeit", 8),
    ("Integrität", 8),
    ("Verfügbarkeit", 8),
    ("Authentizität", 8),
]
SP_ANF, SP_BESCHR, SP_TEIL, SP_SCHUTZ, SP_TEXT = 1, 2, 3, 4, 5
SP_UMS, SP_NACHW, SP_STATUS, SP_MASSN = 6, 7, 8, 9
SP_BEFR, SP_PRUEF, SP_DATUM = 10, 11, 12
SP_SCHWERPUNKT = 13
SP_ALTE_ANF, SP_ALTE_BEZ = 14, 15
SP_ERL, SP_ZIELOBJ, SP_DOKU, SP_ERGEBNIS, SP_GEFAHR, SP_AUFWAND = 16, 17, 18, 19, 20, 21
SP_C, SP_I, SP_V, SP_A = 22, 23, 24, 25

# Zwei getrennt einklappbare Bloecke - Zuordnung und Guidance unabhaengig steuerbar.
# Die Schwerpunkt-Spalte bleibt bewusst dauerhaft sichtbar: sie ist die
# verdichtete Lesefassung der vier Einzelwerte im Guidance-Block.
GRUPPE_ZUORDNUNG = [SP_ALTE_ANF, SP_ALTE_BEZ]
GRUPPE_GUIDANCE = [SP_ERL, SP_ZIELOBJ, SP_DOKU, SP_ERGEBNIS, SP_GEFAHR, SP_AUFWAND,
                   SP_C, SP_I, SP_V, SP_A]

BLATT_DECK = "Deckblatt"
BLATT_PRAKTIKEN = "Praktiken"
BLATT_DASH = "Dashboard"
BLATT_HILF = "Hilfstabelle"
BLATT_GEFAHR = "Gefährdungen"

# Spaltenlage im Blatt "Praktiken" - an EINER Stelle definiert, weil sowohl
# das Blatt selbst als auch die Praktikblaetter (Spiegelung Zielobjekt/
# Zustaendig) und das Dashboard darauf verweisen.
SP_UEB_PRAKTIK = 1
SP_UEB_TITEL = 2
SP_UEB_RELEVANZ = 3
SP_UEB_ZIELOBJEKT = 4
SP_UEB_BEGRUENDUNG = 5
SP_UEB_UMS_PCT = 6
SP_UEB_BEARB_PCT = 7
SP_UEB_STATUS1 = 8  # erste Statuszaehlspalte, danach je Status eine weitere
# Ab hier von der Vokabularlaenge abgeleitet, damit ein zusaetzlicher Status
# nicht stillschweigend die Spaltenlage verschiebt.
SP_UEB_PUNKTE1 = SP_UEB_STATUS1 + len(UMSETZUNGSSTATUS)  # 4 Schutzziel-Punktespalten
SP_UEB_ZUSTAENDIG = SP_UEB_PUNKTE1 + 4
SP_UEB_KOMMENTAR = SP_UEB_ZUSTAENDIG + 1


SCHUTZZIELE = [
    ("vertraulichkeit", "Vertraulichkeit", "C"),
    ("integritaet", "Integrität", "I"),
    ("verfuegbarkeit", "Verfügbarkeit", "V"),
    ("authentizitaet", "Authentizität", "A"),
]


def schwerpunkt(r: Requirement) -> str:
    """
    Verdichtet die vier Schutzziel-Attribute zu einer lesbaren Kurzangabe.

    Die BSI-Skala (documentation/namespaces/security_targets_levels.csv) lautet:
        0 = wirkt nicht oder vernachlaessigbar gering auf dieses Schutzziel
        1 = wirkt auf dieses Schutzziel hin
        2 = wirkt in besonderem Masse; das Schutzziel steht im Zentrum

    Genannt wird deshalb nur, was den Wert 2 traegt - das ist die eigentliche
    Aussage. Der Wert 1 ist im Katalog der Normalfall (haeufigste Kombination
    ist C1/I1/V1 bei 252 Anforderungen) und damit informationsarm.

    Leeres Ergebnis heisst "kein Schutzziel besonders betont", nicht "keine
    Angabe" - rund die Haelfte der Anforderungen faellt darunter. Anforderungen
    ganz ohne Schutzzielangaben (99 Stueck) liefern ebenfalls "".
    """
    zentral = [name for attr, name, _ in SCHUTZZIELE if getattr(r, attr) == 2]
    return ", ".join(zentral)


def schutz_kategorie(r: Requirement) -> str:
    """
    GS++ (2 Stufen + Modalverb) -> 2023er Anforderungskategorie (3 Stufen).

    ANNAHME, nicht vom BSI dokumentiert:
        erhöht                  -> Anforderungen bei erhöhtem Schutzbedarf
        normal-SdT + MUSS       -> Basis-Anforderungen
        normal-SdT + SOLLTE/KANN-> Standard-Anforderungen

    Diese Zuordnung ist der wichtigste fachlich zu bestaetigende Punkt der
    ganzen Umsetzung und steht deshalb auch auf dem Deckblatt.
    """
    if r.schutzbedarfsstufe == "erhöht":
        return "Anforderungen bei erhöhtem Schutzbedarf"
    if r.modalverb == "MUSS":
        return "Basis-Anforderungen"
    return "Standard-Anforderungen"


def _gruppiere(reqs: list[Requirement]) -> list[tuple[Requirement, list[Requirement]]]:
    """Ebene-0-Knoten mit ihren pruefpflichtigen Nachfahren (Tiefensuche-Ordnung)."""
    gruppen: list[tuple[Requirement, list[Requirement]]] = []
    i = 0
    while i < len(reqs):
        kopf = reqs[i]
        if kopf.ebene != 0:  # sollte nicht vorkommen, aber robust bleiben
            i += 1
            continue
        j = i + 1
        while j < len(reqs) and reqs[j].ebene > 0:
            j += 1
        block = reqs[i:j]
        teil = [r for r in block if r.pruefpflichtig]
        gruppen.append((kopf, teil))
        i = j
    return gruppen


def _kopf(ws, headers, zeile) -> None:
    for i, (t, b) in enumerate(headers, start=1):
        c = ws.cell(row=zeile, column=i, value=t)
        c.font = KOPF_FONT
        c.fill = KOPF_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = RAHMEN
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.row_dimensions[zeile].height = 28


def _blatt_praktik(wb: Workbook, praktik_id: str, praktik_titel: str,
                   gruppen: list[tuple[Requirement, list[Requirement]]],
                   praktik_zeile: int | None = None) -> int:
    ws = wb.create_sheet(praktik_id)

    ws["A1"] = "Einstufung gemäß Vorgaben der Sicherheitskonzeption"
    ws["A1"].font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True, color="808080")
    ws["A3"] = praktik_id
    ws["A3"].font = Font(name=FONT, size=design.AKTIV.groesse_deckblatt, bold=True)
    ws["B3"] = praktik_titel
    ws["B3"].font = Font(name=FONT, size=design.AKTIV.groesse_deckblatt, bold=True)

    # Zielobjekt und Zuständig werden zentral im Blatt "Praktiken" gepflegt und
    # hier nur gespiegelt - sonst muesste man dieselbe Angabe zweimal eintragen
    # und beide koennten auseinanderlaufen. Darstellung als berechnete Zelle
    # (grau), damit erkennbar ist: hier nicht eintippen.
    # praktik_zeile wird vom Aufrufer gesetzt; ohne Angabe bleiben Eingabefelder.
    for zeile, label, quell_spalte in (
        (4, "Zielobjekte", SP_UEB_ZIELOBJEKT),
        (5, "Zuständig", SP_UEB_ZUSTAENDIG),
    ):
        ws.cell(row=zeile, column=1, value=label).font = Font(
            name=FONT, size=design.AKTIV.groesse_daten, bold=True)
        c = ws.cell(row=zeile, column=2)
        if praktik_zeile:
            quelle = f"'{BLATT_PRAKTIKEN}'!{get_column_letter(quell_spalte)}{praktik_zeile}"
            c.value = f'=IF({quelle}="","",{quelle})'
            c.fill = ROLLUP_FILL
            c.font = Font(name=FONT, size=design.AKTIV.groesse_daten, italic=True)
        else:
            c.fill = EINGABE_FILL
        c.border = RAHMEN

    _kopf(ws, SPALTEN_2023, KOPFZEILE)
    ws.freeze_panes = ws.cell(row=KOPFZEILE + 1, column=1)
    letzte_spalte = get_column_letter(len(SPALTEN_2023))
    ws.auto_filter.ref = f"A{KOPFZEILE}:{letzte_spalte}{KOPFZEILE}"
    # Gitternetz aus wie in der BSI-Referenzvorlage; rechter Rand als schmale Spalte
    from .excel import blattkosmetik
    blattkosmetik(ws, randspalte_rechts=len(SPALTEN_2023) + 1)

    dv = DataValidation(type="list", formula1='"' + ",".join(UMSETZUNGSSTATUS) + '"',
                        allow_blank=True, showDropDown=False)
    dv.errorTitle = "Ungültiger Status"
    dv.error = "Bitte einen Wert aus der Liste wählen."
    ws.add_data_validation(dv)
    sp_status = get_column_letter(SP_STATUS)

    zeile = KOPFZEILE + 1
    for kopf, teile in gruppen:
        if not teile:
            continue
        sammel_zeile = zeile
        erste_teil = zeile + 1
        letzte_teil = zeile + len(teile)

        # --- Anforderungszeile (Sammelzeile, kein Eingabefeld)
        werte = {
            SP_ANF: kopf.anforderung_id,
            SP_BESCHR: kopf.titel,
            SP_TEIL: None,
            SP_SCHUTZ: schutz_kategorie(kopf),
            SP_TEXT: "Gesamt:",
            SP_STATUS: _rollup_formel(sp_status, erste_teil, letzte_teil),
            SP_SCHWERPUNKT: schwerpunkt(kopf),
            SP_ALTE_ANF: kopf.alte_anforderungen,
            SP_ALTE_BEZ: kopf.alte_beziehung,
            SP_ERL: kopf.erlaeuterung,
            SP_ZIELOBJ: kopf.zielobjekt_kategorien,
            SP_DOKU: kopf.dokumentation,
            SP_ERGEBNIS: kopf.ergebnis,
            SP_GEFAHR: kopf.gefaehrdungen_lang or kopf.gefaehrdungen,
            SP_AUFWAND: kopf.aufwand,
            SP_C: kopf.vertraulichkeit,
            SP_I: kopf.integritaet,
            SP_V: kopf.verfuegbarkeit,
            SP_A: kopf.authentizitaet,
        }
        for sp in range(1, len(SPALTEN_2023) + 1):
            c = ws.cell(row=sammel_zeile, column=sp, value=werte.get(sp))
            c.border = RAHMEN
            c.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=sp in (SP_ANF, SP_BESCHR))
            c.alignment = Alignment(vertical="top",
                                    wrap_text=sp in (SP_BESCHR, SP_TEXT, SP_ALTE_ANF, SP_ERL, SP_ERGEBNIS, SP_GEFAHR))
        ws.cell(row=sammel_zeile, column=SP_STATUS).fill = ROLLUP_FILL
        ws.cell(row=sammel_zeile, column=SP_STATUS).font = Font(name=FONT, size=design.AKTIV.groesse_daten, italic=True)

        # --- Teilanforderungszeilen (Eingabe)
        for versatz, t in enumerate(teile, start=1):
            z = sammel_zeile + versatz
            werte = {
                SP_ANF: kopf.anforderung_id,
                SP_BESCHR: kopf.titel,
                SP_TEIL: t.anforderung_id,
                SP_SCHUTZ: schutz_kategorie(t),
                SP_TEXT: t.anforderungstext,
                SP_SCHWERPUNKT: schwerpunkt(t),
                SP_ALTE_ANF: t.alte_anforderungen,
                SP_ALTE_BEZ: t.alte_beziehung,
                SP_ERL: t.erlaeuterung,
                SP_ZIELOBJ: t.zielobjekt_kategorien,
                SP_DOKU: t.dokumentation,
                SP_ERGEBNIS: t.ergebnis,
                SP_GEFAHR: t.gefaehrdungen_lang or t.gefaehrdungen,
                SP_AUFWAND: t.aufwand,
                SP_C: t.vertraulichkeit,
                SP_I: t.integritaet,
                SP_V: t.verfuegbarkeit,
                SP_A: t.authentizitaet,
            }
            for sp in range(1, len(SPALTEN_2023) + 1):
                c = ws.cell(row=z, column=sp, value=werte.get(sp))
                c.border = RAHMEN
                c.font = ZELL_FONT
                c.alignment = Alignment(vertical="top",
                                        wrap_text=sp in (SP_BESCHR, SP_TEXT, SP_ALTE_ANF, SP_ERL, SP_ERGEBNIS, SP_GEFAHR))
            for sp in (SP_UMS, SP_NACHW, SP_STATUS, SP_MASSN, SP_BEFR, SP_PRUEF, SP_DATUM):
                ws.cell(row=z, column=sp).fill = EINGABE_FILL
            ws.cell(row=z, column=SP_STATUS, value="unbearbeitet")
            dv.add(f"{sp_status}{z}")
            ws.cell(row=z, column=SP_DATUM).number_format = "DD.MM.YYYY"

        zeile = letzte_teil + 1

    # Zwei unabhaengig einklappbare Spaltenbloecke.
    # Alle Spalten starten SICHTBAR - so sieht man beim ersten Oeffnen den
    # vollen Umfang der Vorlage. Wer es kompakter mag, klappt die Bloecke
    # ueber die '-' Schaltflaechen oberhalb der Spaltenkoepfe ein.
    ws.sheet_properties.outlinePr.summaryRight = True
    for sp in GRUPPE_ZUORDNUNG + GRUPPE_GUIDANCE:
        ws.column_dimensions[get_column_letter(sp)].outlineLevel = 1

    from .excel import status_formatierung
    status_formatierung(ws, get_column_letter(SP_STATUS), KOPFZEILE + 1, zeile - 1)

    return zeile - KOPFZEILE - 1


def _blatt_gefaehrdungen(wb: Workbook, tabelle: dict[str, tuple[str, str]],
                         verwendete: list[str]) -> None:
    """
    Nachschlageblatt mit den vollstaendigen Definitionen der elementaren
    Gefaehrdungen (BSI G 0.x).

    Aufgefuehrt werden nur die im Katalog tatsaechlich referenzierten
    Gefaehrdungen, nicht der komplette BSI-Bestand - das haelt das Blatt
    an den Inhalt der Vorlage gebunden.
    """
    ws = wb.create_sheet(BLATT_GEFAHR)
    kopf = [("Gefährdung", 12), ("Begriff", 46), ("Definition", 130)]
    _kopf(ws, kopf, 1)
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = "A1:C1"

    from .excel import blattkosmetik
    blattkosmetik(ws)

    for i, gid in enumerate(verwendete, start=2):
        begriff, definition = tabelle.get(gid, ("(nicht im BSI-Namespace gefunden)", ""))
        for sp, wert in ((1, gid), (2, begriff), (3, definition)):
            c = ws.cell(row=i, column=sp, value=wert)
            c.font = ZELL_FONT if sp == 3 else Font(
                name=FONT, size=design.AKTIV.groesse_daten, bold=sp == 1)
            c.alignment = Alignment(vertical="top", wrap_text=sp in (2, 3))
        # Zeilenhoehe an die Textlaenge anpassen, gedeckelt
        ws.row_dimensions[i].height = min(240, 15 + 11 * (len(definition) // 115))

    hinweis = ws.cell(row=len(verwendete) + 3, column=1,
                      value="Quelle: BSI, documentation/namespaces/basethreats.csv. "
                            "Aufgeführt sind nur die in diesem Katalog referenzierten "
                            "Gefährdungen.")
    hinweis.font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True,
                        color=design.AKTIV.sekundaertext)


def _blatt_praktiken_uebersicht(wb: Workbook, praktiken: dict[str, str],
                                zeilen_je_blatt: dict[str, int]) -> None:
    """Gegenstück zum Blatt 'Modellierung' der BSI-Vorlage."""
    ws = wb.create_sheet(BLATT_PRAKTIKEN, 1)
    header = [("Praktik", 12), ("Titel", 34), ("Relevanz", 14), ("Zielobjekt", 22),
              ("Begründung", 30), ("Umsetzungsstand (%)", 18), ("Bearbeitungsstand (%)", 20)]
    header += [(s, 13) for s in UMSETZUNGSSTATUS]
    header += [(f"Punkte {n}", 13) for _, n, _ in SCHUTZZIELE]
    header += [("Zuständig", 18), ("Kommentar", 30)]
    _kopf(ws, header, 4)
    ws.freeze_panes = ws["A5"]

    dv = DataValidation(type="list", formula1='"' + ",".join(RELEVANZ) + '"', allow_blank=True,
                        showDropDown=False)
    ws.add_data_validation(dv)

    SP_REL = SP_UEB_RELEVANZ
    SP_UMS_PCT, SP_BEARB_PCT = SP_UEB_UMS_PCT, SP_UEB_BEARB_PCT
    SP_ST1 = SP_UEB_STATUS1
    SP_PKT1 = SP_UEB_PUNKTE1
    SP_ZUST = SP_UEB_ZUSTAENDIG

    for i, (pid, titel) in enumerate(sorted(praktiken.items()), start=5):
        ws.cell(row=i, column=1, value=pid).font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
        ws.cell(row=i, column=2, value=titel).font = ZELL_FONT
        for sp in (SP_REL, 4, 5, SP_ZUST, SP_ZUST + 1):
            c = ws.cell(row=i, column=sp)
            c.fill = EINGABE_FILL
            c.border = RAHMEN
        dv.add(f"C{i}")

        # Nur Teilanforderungszeilen zaehlen. Sammelzeilen haben Spalte C leer
        # und in H ein Rollup-Ergebnis - wuerden sie mitgezaehlt, taucht jede
        # Anforderung doppelt auf und der Bearbeitungsstand ist zu hoch.
        n = zeilen_je_blatt.get(pid, 0)
        h = f"'{pid}'!$H${KOPFZEILE + 1}:$H${KOPFZEILE + n}"
        c = f"'{pid}'!$C${KOPFZEILE + 1}:$C${KOPFZEILE + n}"
        teil = f'{c},"<>"'
        for j, st in enumerate(UMSETZUNGSSTATUS, start=SP_ST1):
            ws.cell(row=i, column=j,
                    value=f'=IF($C{i}="Ja",COUNTIFS({teil},{h},"{st}"),"")')
        erl = (f'COUNTIFS({teil},{h},"ja")+COUNTIFS({teil},{h},"entbehrlich")'
               f'+COUNTIFS({teil},{h},"verwiesen")')
        ges = f"COUNTA({c})"
        ws.cell(row=i, column=SP_UMS_PCT,
                value=f'=IF($C{i}="Ja",IFERROR(({erl})/{ges},0),"")')
        bearb = f'{ges}-COUNTIFS({teil},{h},"unbearbeitet")'
        ws.cell(row=i, column=SP_BEARB_PCT,
                value=f'=IF($C{i}="Ja",IFERROR(({bearb})/{ges},0),"")')
        # Punkte je Schutzziel: Summe der Schutzzielwerte (0-2) ueber alle
        # Teilanforderungen mit Status "ja". Nur "ja" - "entbehrlich" und
        # "verwiesen" zaehlen hier bewusst NICHT mit, weil der Schutzbeitrag
        # dort nicht tatsaechlich erbracht wird (Festlegung Sven, 07.08.).
        # Sammelzeilen sind ueber das Kriterium Teilanforderung<>"" ausgeschlossen,
        # sonst wuerde jede Anforderung doppelt gewichtet.
        for versatz, (attr, _, _) in enumerate(SCHUTZZIELE):
            quelle = get_column_letter(SP_C + versatz)
            bereich = f"'{pid}'!${quelle}${KOPFZEILE + 1}:${quelle}${KOPFZEILE + n}"
            ws.cell(row=i, column=SP_PKT1 + versatz,
                    value=f'=IF($C{i}="Ja",SUMIFS({bereich},{teil},{h},"ja"),"")')

        for sp in (SP_UMS_PCT, SP_BEARB_PCT):
            ws.cell(row=i, column=sp).number_format = "0.0%"
        for sp in range(SP_UMS_PCT, SP_ZUST):
            c = ws.cell(row=i, column=sp)
            c.font = ZELL_FONT
            c.border = RAHMEN

    hinweis = ws.cell(row=len(praktiken) + 7, column=2,
                      value='Nur Praktiken mit Relevanz = "Ja" werden im Dashboard gezählt. '
                            'Die Zähler beziehen sich auf Teilanforderungen, nicht auf '
                            'Sammelzeilen.')
    hinweis.font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True, color="808080")


def _blatt_dashboard(wb: Workbook, anzahl_praktiken: int) -> None:
    ws = wb.create_sheet(BLATT_DASH, 1)
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 14
    letzte = anzahl_praktiken + 4
    p = f"{BLATT_PRAKTIKEN}!"
    st_spalte = {s: get_column_letter(8 + i) for i, s in enumerate(UMSETZUNGSSTATUS)}

    zeilen = [
        ("Status Bearbeitung", "Anzahl"),
        ("Relevante Praktiken", f'=COUNTIF({p}$C$5:$C${letzte},"Ja")'),
        ("Abgeschlossene Praktiken",
         f'=COUNTIFS({p}$C$5:$C${letzte},"Ja",{p}$G$5:$G${letzte},1)'),
        ("Begonnene Praktiken",
         f'=COUNTIFS({p}$C$5:$C${letzte},"Ja",{p}$G$5:$G${letzte},">0",'
         f'{p}$G$5:$G${letzte},"<1")'),
        ("", ""),
        ("Status Umsetzung", "Anzahl"),
    ]
    for st in UMSETZUNGSSTATUS:
        col = st_spalte[st]
        zeilen.append((f"Teilanforderungen: {st}",
                       f'=SUMIFS({p}${col}$5:${col}${letzte},{p}$C$5:$C${letzte},"Ja")'))

    # Erreichte Punkte je Schutzziel, ueber alle relevanten Praktiken.
    # Quelle sind die Punktespalten im Blatt "Praktiken", die dort bereits
    # auf Status "ja" und Teilanforderungszeilen gefiltert sind.
    zeilen.append(("", ""))
    zeilen.append(("Erreichte Punkte je Schutzziel", "Punkte"))
    for versatz, (_, name, _) in enumerate(SCHUTZZIELE):
        col = get_column_letter(SP_UEB_PUNKTE1 + versatz)
        zeilen.append((name,
                       f'=SUMIFS({p}${col}$5:${col}${letzte},{p}$C$5:$C${letzte},"Ja")'))
    zeilen.append(("", ""))

    for i, (a, b) in enumerate(zeilen, start=2):
        ca = ws.cell(row=i, column=2, value=a)
        cb = ws.cell(row=i, column=3, value=b)
        fett = b in ("Anzahl",)
        ca.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=fett)
        cb.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=fett)
        if fett:
            ca.fill = KOPF_FILL
            cb.fill = KOPF_FILL
            ca.font = KOPF_FONT
            cb.font = KOPF_FONT

    hinweis = ws.cell(row=len(zeilen) + 3, column=2,
                      value='Punkte = Summe der Schutzzielwerte (0–2) aller '
                            'Teilanforderungen mit Status "ja". Entbehrliche und '
                            'verwiesene Anforderungen zählen nicht mit.')
    hinweis.font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True,
                        color=design.AKTIV.sekundaertext)


def _blatt_hilfstabelle(wb: Workbook) -> None:
    ws = wb.create_sheet(BLATT_HILF)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["C"].width = 30
    ws.cell(row=1, column=1, value="Relevanz").font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
    for i, v in enumerate(RELEVANZ, start=2):
        ws.cell(row=i, column=1, value=v).font = ZELL_FONT
    ws.cell(row=1, column=3, value="Status").font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
    for i, v in enumerate(UMSETZUNGSSTATUS, start=2):
        ws.cell(row=i, column=3, value=v).font = ZELL_FONT
    ws.sheet_state = "hidden"


def _kein_commit(snap: CatalogSnapshot) -> str:
    """Ehrliche Beschriftung statt pauschalem 'Offline'."""
    if snap.quelle_url.startswith("file://"):
        return "n/a — lokale Quelldatei, SHA-256 ist Versionsanker"
    return "n/a — GitHub-API nicht abgefragt, SHA-256 ist Versionsanker"


def _blatt_deckblatt(wb: Workbook, snap: CatalogSnapshot, anzahl_teil: int,
                     anzahl_anf: int) -> None:
    ws = wb.create_sheet(BLATT_DECK, 0)
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 62

    ws["B1"] = "Einstufung gemäß Vorgaben der Sicherheitskonzeption"
    ws["B1"].fill = KOPF_FILL
    ws["B1"].font = KOPF_FONT

    zeilen = [
        ("Dokumententitel", "Praktiken und Grundschutz++-Checks"),
        ("Layout", "BSI-Vorlage A.3.4, Kompendium-Edition 2023"),
        ("Katalog", "Grundschutz++ (OSCAL), Anwenderkatalog des BSI"),
        ("Katalog-Version", snap.katalog_version),
        ("Git-Commit", snap.commit_sha or _kein_commit(snap)),
        ("SHA-256 der Quelldatei", snap.sha256_quelle),
        ("Erzeugt am (UTC)", snap.abgerufen_am),
        ("Werkzeug", f"gspp-tool {snap.tool_version}"),
        ("", ""),
        ("Anforderungszeilen", anzahl_anf),
        ("Teilanforderungen (bewertbar)", anzahl_teil),
        ("", ""),
        ("Zuordnung zu IT-Grundschutz 2023", ""),
        ("Quelle", "Offizielles BSI-Mapping (control_layer/Mappings/IT-GS2023-zu-GSpp)"),
        ("Abdeckung", "Unvollstaendig, laut BSI noch in Pilotierung - nur ca. 30% der "
                      "GS++-Anforderungen haben aktuell eine Zuordnung. Leere Spalten "
                      "'Alte Anforderung' bedeuten 'noch nicht kartiert', nicht "
                      "'keine Beziehung'."),
        ("", ""),
        ("Spalten ein-/ausblenden", ""),
        ("Bedienung", "Alle 25 Spalten sind beim Öffnen sichtbar. Über den "
                      "Spaltenköpfen stehen kleine +/− Schaltflächen, mit denen sich "
                      "zwei Blöcke unabhängig einklappen lassen: "
                      "'Zuordnung IT-GS 2023' (Spalten N–O) und 'Guidance' "
                      "(Spalten P–Y: Erläuterung, Zielobjekte, Dokumentation, "
                      "Ergebnis, Gefährdungen, Aufwand, vier Schutzziele). "
                      "Keine Makros nötig."),
        ("Schwerpunkt (Schutzziel)", "Nennt nur Schutzziele mit Wert 2 – also jene, die "
                                     "laut BSI im Zentrum der Anforderung stehen. Leer "
                                     "heißt: kein Schutzziel besonders betont (gilt für "
                                     "rund die Hälfte der Anforderungen). Die Einzelwerte "
                                     "0/1/2 stehen im Guidance-Block."),
        ("", ""),
        ("Offene fachliche Annahme", ""),
        ("Spalte 'Schutz'",
         "GS++ kennt zwei Schutzbedarfsstufen, die 2023er-Vorlage drei "
         "Anforderungskategorien. Abgebildet wird: erhöht -> erhöhter Schutzbedarf; "
         "normal-SdT + MUSS -> Basis; normal-SdT + SOLLTE/KANN -> Standard. "
         "Diese Zuordnung ist nicht vom BSI dokumentiert und fachlich zu bestätigen."),
    ]
    for i, (k, v) in enumerate(zeilen, start=3):
        a = ws.cell(row=i, column=2, value=k)
        b = ws.cell(row=i, column=3, value=v if v != "" else None)
        a.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
        b.font = ZELL_FONT
        b.alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(v, str) and len(v) > 100:
            ws.row_dimensions[i].height = 20 + 14 * (len(v) // 70)

    # Steuerzellen fuer Makro-Bedienung sind bewusst NICHT enthalten:
    # Ohne aktivierte Makros waeren es zwei Zellen ohne Wirkung - das
    # verwirrt mehr als es hilft. Die Spaltenbloecke werden ueber die
    # +/- Schaltflaechen der Spaltengruppierung bedient.


def _steuerzellen(ws, startzeile: int) -> None:
    """
    Ja/Nein-Auswahlzellen zur Spaltensteuerung, als benannte Bereiche.

    Ohne aktivierte Makros sind das schlicht zwei beschriftete Zellen ohne
    Wirkung - deshalb steht der Hinweis auf die +/- Schaltflaechen direkt
    daneben. Die Vorlage bleibt damit in jedem Fall bedienbar.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import quote_sheetname, absolute_coordinate

    ws.cell(row=startzeile, column=2, value="Ansicht steuern (benötigt Makros)").font = \
        Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True, color=design.AKTIV.akzent)

    dv = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=False,
                        showDropDown=False)
    ws.add_data_validation(dv)

    eintraege = [
        ("Guidance-Spalten anzeigen", "Nein", "GUIDANCE_SICHTBAR"),
        ("Zuordnung IT-GS 2023 anzeigen", "Ja", "ZUORDNUNG_SICHTBAR"),
    ]
    for versatz, (label, vorgabe, name) in enumerate(eintraege, start=1):
        z = startzeile + versatz
        ws.cell(row=z, column=2, value=label).font = ZELL_FONT
        c = ws.cell(row=z, column=3, value=vorgabe)
        c.font = ZELL_FONT
        c.fill = EINGABE_FILL
        dv.add(c)
        ws.parent.defined_names[name] = DefinedName(
            name, attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(c.coordinate)}"
        )

    hinweis = ws.cell(row=startzeile + 3, column=2,
                      value="Ohne aktivierte Makros wirkungslos — nutzen Sie dann die "
                            "+/− Schaltflächen über den Spaltenköpfen der Praktikblätter.")
    hinweis.font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True,
                        color=design.AKTIV.sekundaertext)


def erzeuge_mappe_2023(snap: CatalogSnapshot, ziel: Path,
                       reqs: list[Requirement] | None = None,
                       gefaehrdungen: dict[str, tuple[str, str]] | None = None) -> Path:
    reqs = reqs if reqs is not None else snap.requirements
    wb = Workbook()
    wb.remove(wb.active)

    praktiken: dict[str, str] = {}
    for r in reqs:
        praktiken.setdefault(r.praktik_id, r.praktik_titel)

    nach_praktik: dict[str, list[Requirement]] = {}
    for r in reqs:
        nach_praktik.setdefault(r.praktik_id, []).append(r)

    zeilen_je_blatt: dict[str, int] = {}
    anzahl_anf = anzahl_teil = 0
    # Zeilennummer im Blatt "Praktiken": dort beginnen die Daten in Zeile 5,
    # sortiert nach Praktik-ID. Muss zur Schleife dort identisch sein.
    for idx, pid in enumerate(sorted(praktiken), start=5):
        gruppen = _gruppiere(nach_praktik[pid])
        gruppen = [(k, t) for k, t in gruppen if t]
        n = _blatt_praktik(wb, pid, praktiken[pid], gruppen, praktik_zeile=idx)
        zeilen_je_blatt[pid] = n
        anzahl_anf += len(gruppen)
        anzahl_teil += sum(len(t) for _, t in gruppen)

    if gefaehrdungen:
        from .threats import verwendete_ids
        verwendete = verwendete_ids(reqs)
        if verwendete:
            _blatt_gefaehrdungen(wb, gefaehrdungen, verwendete)
            log.info("Nachschlageblatt Gefährdungen: %d Einträge", len(verwendete))

    _blatt_praktiken_uebersicht(wb, praktiken, zeilen_je_blatt)
    _blatt_dashboard(wb, len(praktiken))
    _blatt_hilfstabelle(wb)
    _blatt_deckblatt(wb, snap, anzahl_teil, anzahl_anf)
    _blatt_metadaten(wb, snap)

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)
    log.info("2023-Schema geschrieben: %s (%d Anforderungszeilen, %d Teilanforderungen, "
             "%d Praktik-Blätter)", ziel, anzahl_anf, anzahl_teil, len(praktiken))
    return ziel
