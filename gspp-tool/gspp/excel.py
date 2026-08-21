"""
Render-Layer: Anforderungen -> Excel-Vorlagen.

Zwei Betriebsarten:

  TEMPLATE  (bevorzugt)  Eine leergeraeumte Kopie der bestehenden 2023er-Vorlage
                         wird geoeffnet und nur befuellt. Corporate Design,
                         Spaltenbreiten, Kopfzeilen und Dropdowns bleiben so
                         erhalten. Spalten werden ueber die Kopfzeile gemappt,
                         nicht ueber feste Indizes.

  GENERATE  (Fallback)   Arbeitsmappe wird komplett erzeugt. Nuetzlich fuer den
                         ersten Wurf und fuer Regressionstests.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .models import CatalogSnapshot, Requirement
from .parser import filter_mit_ahnen

log = logging.getLogger(__name__)

from . import design

FONT = design.AKTIV.schrift

BLATT_META = "00_Metadaten"
BLATT_UEBERSICHT = "01_Uebersicht"
BLATT_ANF = "02_Anforderungen"
BLATT_CHECK = "03_Grundschutz-Check"
BLATT_LEGENDE = "04_Legende"

# Vokabular 1:1 aus der BSI-Vorlage A.3.4 (Blatt "Hilfstabelle", Spalte C)
# plus "verwiesen", das dort nur in den Rollup-Formeln vorkommt.
UMSETZUNGSSTATUS = ["ja", "nein", "teilweise", "entbehrlich", "verwiesen",
                    "zu klären", "unbearbeitet"]
STATUS_UNVOLLSTAENDIG = "Bitte Teilanforderungen ausfüllen"

# Relevanz-Vokabular des Modellierungsblatts
RELEVANZ = ["Ja", "Nein", "Nein, in anderem SiKo betrachtet"]

# Spaltenkopf -> Attribut auf Requirement. Reihenfolge = Spaltenreihenfolge.
SPALTEN: list[tuple[str, str, int]] = [
    ("Praktik", "praktik_id", 10),
    ("Praktik-Bezeichnung", "praktik_titel", 26),
    ("Thema", "thema_id", 10),
    ("Thema-Bezeichnung", "thema_titel", 26),
    ("Anforderungs-ID", "anforderung_id", 16),
    ("Ebene", "ebene", 7),
    ("Knotentyp", "knotentyp", 26),
    ("Uebergeordnet", "parent_id", 16),
    ("Pfad", "pfad", 30),
    ("Titel", "titel", 42),
    ("Anforderung", "anforderungstext", 70),
    ("Modalverb", "modalverb", 11),
    ("Schutzbedarfsstufe", "schutzbedarfsstufe", 18),
    ("Aufwand", "aufwand", 9),
    ("Zielobjekt-Kategorien", "zielobjekt_kategorien", 24),
    ("Dokumentation", "dokumentation", 24),
    ("Ergebnis", "ergebnis", 40),
    ("Vertraulichkeit", "vertraulichkeit", 14),
    ("Integritaet", "integritaet", 12),
    ("Verfuegbarkeit", "verfuegbarkeit", 14),
    ("Authentizitaet", "authentizitaet", 14),
    ("Gefaehrdungen", "gefaehrdungen", 20),
    ("Tags", "tags", 22),
    ("Verweise (related)", "verweise_related", 22),
    ("Verweise (required)", "verweise_required", 22),
    ("Erlaeuterung", "erlaeuterung", 80),
    ("Alte Anforderung (IT-GS 2023)", "alte_anforderungen", 40),
    ("Beziehung (IT-GS 2023)", "alte_beziehung", 16),
]

# Spalten, die standardmaessig sichtbar aber einklappbar sind (Excel-Gruppierung,
# +/- Klick - kein VBA noetig). Jede Menge wird als eigene, unabhaengige Gruppe
# aus zusammenhaengenden Spalten behandelt.
EINKLAPPBARE_GRUPPEN: list[set[str]] = [
    {"erlaeuterung"},
    {"alte_anforderungen", "alte_beziehung"},
]

# Spaltenreihenfolge des Check-Blatts - einzige Wahrheit fuer Blatt und Formeln.
CHECK_KERN = ["anforderung_id", "praktik_id", "thema_id", "ebene", "knotentyp",
              "__pruefpflichtig", "titel", "anforderungstext", "modalverb",
              "schutzbedarfsstufe", "aufwand", "zielobjekt_kategorien"]


def check_spalte(attr: str) -> str:
    """Spaltenbuchstabe im Check-Blatt. Verhindert per-Hand gezaehlte Offsets."""
    from openpyxl.utils import get_column_letter as _g
    return _g(CHECK_KERN.index(attr) + 1)


CHECK_ZUSATZSPALTEN = [
    ("Umsetzungsstatus", 18),
    ("Umsetzung bis", 14),
    ("Verantwortlich", 20),
    ("Begruendung / Bemerkung", 45),
    ("Nachweis / Referenz", 30),
]

KOPF_FILL = design.AKTIV.kopf_fill()
KOPF_FONT = design.AKTIV.kopf_font()
ZELL_FONT = design.AKTIV.zell_font()
EINGABE_FILL = design.AKTIV.eingabe_fill()
ROLLUP_FILL = design.AKTIV.berechnet_fill()
# Referenz A.1 arbeitet ohne Zellrahmen. RAHMEN bleibt als No-Op erhalten,
# damit bestehende Aufrufstellen unveraendert bleiben koennen.
DUENN = Side(style="thin", color=design.AKTIV.trennlinie)
RAHMEN = (Border(left=DUENN, right=DUENN, top=DUENN, bottom=DUENN)
          if design.AKTIV.zellrahmen else Border())


def blattkosmetik(ws: Worksheet, randspalte_rechts: int | None = None) -> None:
    """
    Gitternetz aus, schmale Randspalten - wie in der BSI-Referenzvorlage A.1.

    randspalte_rechts: 1-indizierte Spalte, die als rechter Rand schmal wird.
    """
    ws.sheet_view.showGridLines = design.AKTIV.gitternetz
    if randspalte_rechts:
        ws.column_dimensions[get_column_letter(randspalte_rechts)].width = \
            design.AKTIV.randspaltenbreite


# --------------------------------------------------------------------- Helfer


def _spalten_gruppieren(ws: Worksheet, attribute: list[str], gruppen: list[set[str]]) -> None:
    """
    Setzt Excel-Spaltengruppierung (+/- Klick) fuer die angegebenen Attributmengen.

    attribute: Reihenfolge der Spalten wie tatsaechlich geschrieben (1-indiziert ueber Position)
    gruppen: Mengen von Attributnamen, die jeweils eine eigene Gruppe bilden
    """
    ws.sheet_properties.outlinePr.summaryRight = True
    for gruppe in gruppen:
        for i, attr in enumerate(attribute, start=1):
            if attr in gruppe:
                ws.column_dimensions[get_column_letter(i)].outlineLevel = 1


def status_formatierung(ws: Worksheet, spalte: str, von: int, bis: int) -> None:
    """
    Ampelfarben fuer die Statusspalte per bedingter Formatierung.

    Bewusst als Regel statt als feste Zellfuellung: bedingte Formatierung
    reagiert auf den ANGEZEIGTEN Wert. Damit faerben sich Sammelzeilen, deren
    Status aus einer Rollup-Formel stammt, genauso ein wie manuell
    ausgefuellte Teilanforderungen - ohne dass der Code zwischen beiden
    Zeilenarten unterscheiden muesste.

    'unbearbeitet' bekommt bewusst keine Regel: der Normalzustand soll nicht
    auffallen, sonst ist beim ersten Oeffnen das ganze Blatt bunt.
    """
    if bis < von:
        return
    bereich = f"{spalte}{von}:{spalte}{bis}"
    for status, (fuellung, schrift) in design.AKTIV.status_farben.items():
        regel = CellIsRule(
            operator="equal",
            formula=[f'"{status}"'],
            fill=PatternFill("solid", start_color=fuellung, end_color=fuellung),
            font=Font(name=design.AKTIV.schrift, size=design.AKTIV.groesse_daten,
                      color=schrift),
        )
        ws.conditional_formatting.add(bereich, regel)


def _kopfzeile(ws: Worksheet, headers: list[tuple[str, int]], zeile: int = 1) -> None:
    for i, (titel, breite) in enumerate(headers, start=1):
        c = ws.cell(row=zeile, column=i, value=titel)
        c.font = KOPF_FONT
        c.fill = KOPF_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = RAHMEN
        ws.column_dimensions[get_column_letter(i)].width = breite
    ws.row_dimensions[zeile].height = 30
    ws.freeze_panes = ws.cell(row=zeile + 1, column=1)
    ws.auto_filter.ref = f"A{zeile}:{get_column_letter(len(headers))}{zeile}"


def _schreibe_zeilen(ws: Worksheet, reqs: list[Requirement], startzeile: int,
                     attribute: list[str]) -> None:
    for zi, r in enumerate(reqs, start=startzeile):
        for si, attr in enumerate(attribute, start=1):
            wert = getattr(r, attr)
            c = ws.cell(row=zi, column=si, value=wert)
            c.font = ZELL_FONT
            c.alignment = Alignment(vertical="top", wrap_text=attr in
                                    {"anforderungstext", "erlaeuterung", "titel", "ergebnis"})
            c.border = RAHMEN


# ------------------------------------------------------------------- Blaetter


def _blatt_metadaten(wb: Workbook, snap: CatalogSnapshot) -> None:
    ws = wb.create_sheet(BLATT_META, 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 78

    zeilen = [
        ("Herkunftsnachweis", ""),
        ("Quelle", snap.quelle_url),
        ("Git-Commit (SHA)", snap.commit_sha or (
            "n/a — lokale Quelldatei, SHA-256 ist Versionsanker"
            if snap.quelle_url.startswith("file://")
            else "n/a — GitHub-API nicht abgefragt, SHA-256 ist Versionsanker")),
        ("Commit-Datum", snap.commit_datum or "n/a"),
        ("SHA-256 der Quelldatei", snap.sha256_quelle),
        ("Abgerufen am (UTC)", snap.abgerufen_am),
        ("", ""),
        ("Katalog", ""),
        ("Katalog-UUID", snap.katalog_uuid),
        ("Katalog-Version", snap.katalog_version),
        ("Zuletzt geaendert (BSI)", snap.katalog_last_modified),
        ("OSCAL-Version", snap.oscal_version),
        ("Anzahl Anforderungen", len(snap.requirements)),
        ("", ""),
        ("Erzeugung", ""),
        ("Werkzeug-Version", snap.tool_version),
    ]
    for i, (k, v) in enumerate(zeilen, start=1):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=not v or k in ("Herkunftsnachweis", "Katalog", "Erzeugung"))
        b.font = ZELL_FONT
    ws["A1"].fill = KOPF_FILL
    ws["A1"].font = KOPF_FONT


def _blatt_anforderungen(wb: Workbook, reqs: list[Requirement]) -> None:
    ws = wb.create_sheet(BLATT_ANF)
    _kopfzeile(ws, [(t, b) for t, _, b in SPALTEN])
    _schreibe_zeilen(ws, reqs, 2, [a for _, a, _ in SPALTEN])
    _spalten_gruppieren(ws, [a for _, a, _ in SPALTEN], EINKLAPPBARE_GRUPPEN)


def _blatt_check(wb: Workbook, reqs: list[Requirement]) -> None:
    """
    Arbeitsblatt fuer den Grundschutz-Check.

    Umgang mit Verschachtelung:
      * Titel wird je Ebene eingerueckt, Excel-Gliederungsebenen erlauben
        Ein-/Ausklappen ganzer Anforderungsbaeume.
      * Sammelanforderungen bekommen KEIN Eingabefeld. Ihr Status wird per
        Formel aus den Teilanforderungen aggregiert (grau statt gelb).
        Sonst bewertet man dieselbe Leistung zweimal und der Erfuellungsgrad
        wird geschoent.
      * Die Spalte "Pruefpflichtig" ist das Kriterium, ueber das die Uebersicht
        zaehlt - nur so stimmen Anzahl und Aufwandssumme.
    """
    ws = wb.create_sheet(BLATT_CHECK)

    kern = [a for a in CHECK_KERN if a != "__pruefpflichtig"]
    kopf_map = {a: t for t, a, _ in SPALTEN}
    breit_map = {a: b for _, a, b in SPALTEN}
    header = [("Pruefpflichtig", 13) if a == "__pruefpflichtig"
              else (kopf_map[a], breit_map[a]) for a in CHECK_KERN]
    header += CHECK_ZUSATZSPALTEN
    _kopfzeile(ws, header)

    SP_PFLICHT = CHECK_KERN.index("__pruefpflichtig") + 1
    SP_TITEL = CHECK_KERN.index("titel") + 1
    erste_eingabe = len(CHECK_KERN) + 1
    SP_STATUS = erste_eingabe
    letzte = len(reqs) + 1

    # Zeilennummer je Anforderungs-ID (fuer die Rollup-Bereiche)
    zeile_von = {r.anforderung_id: i for i, r in enumerate(reqs, start=2)}

    for zi, r in enumerate(reqs, start=2):
        werte = [
            ("ja" if r.pruefpflichtig else "nein") if a == "__pruefpflichtig"
            else getattr(r, a)
            for a in CHECK_KERN
        ]
        # Einrueckung ueber den Titel, damit der Baum lesbar bleibt
        werte[SP_TITEL - 1] = ("    " * r.ebene) + str(werte[SP_TITEL - 1])
        for si, v in enumerate(werte, start=1):
            c = ws.cell(row=zi, column=si, value=v)
            c.font = ZELL_FONT
            c.alignment = Alignment(vertical="top",
                                    wrap_text=si in (SP_TITEL, SP_TITEL + 1))
            c.border = RAHMEN
        if r.ebene:
            ws.row_dimensions[zi].outlineLevel = min(r.ebene, 7)
        if not r.pruefpflichtig:
            for si in range(1, SP_PFLICHT + 1):
                ws.cell(row=zi, column=si).font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)

    ws.sheet_properties.outlinePr.summaryBelow = False

    # ---- Eingabefelder bzw. Rollup je Zeile
    dv = DataValidation(type="list", formula1='"' + ",".join(UMSETZUNGSSTATUS) + '"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Bitte einen Wert aus der Liste waehlen."
    dv.errorTitle = "Ungueltiger Umsetzungsstatus"
    ws.add_data_validation(dv)
    sp_status = get_column_letter(SP_STATUS)

    for zi, r in enumerate(reqs, start=2):
        if r.pruefpflichtig:
            for si in range(erste_eingabe, erste_eingabe + len(CHECK_ZUSATZSPALTEN)):
                c = ws.cell(row=zi, column=si)
                c.fill = EINGABE_FILL
                c.font = ZELL_FONT
                c.border = RAHMEN
            dv.add(f"{sp_status}{zi}")
            ws.cell(row=zi, column=erste_eingabe + 1).number_format = "DD.MM.YYYY"
        else:
            von, bis = _nachfahren_bereich(reqs, zi, r)
            c = ws.cell(row=zi, column=SP_STATUS,
                        value=_rollup_formel(sp_status, von, bis))
            c.fill = ROLLUP_FILL
            c.font = Font(name=FONT, size=design.AKTIV.groesse_daten, italic=True)
            c.border = RAHMEN
            ws.cell(row=zi, column=erste_eingabe + 3,
                    value="aggregiert aus Teilanforderungen").font = Font(
                        name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True, color="808080")

    ws.cell(row=letzte + 2, column=1,
            value="Graue Statuszellen werden berechnet - nur gelbe Zellen ausfuellen."
            ).font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True, color="808080")


def _nachfahren_bereich(reqs: list[Requirement], zeile: int, r: Requirement) -> tuple[int, int]:
    """Nachfahren liegen dank Tiefensuche zusammenhaengend direkt unter dem Knoten."""
    idx = zeile - 2
    ende = idx + 1
    while ende < len(reqs) and reqs[ende].ebene > r.ebene:
        ende += 1
    return zeile + 1, ende + 1


def _rollup_formel(spalte: str, von: int, bis: int) -> str:
    """
    Aggregation eines Sammelknotens - Semantik 1:1 aus der BSI-Vorlage A.3.4
    (dort z.B. APP.3.3!H9). Reihenfolge der Prüfungen ist bedeutsam:

      leer oder alles unbearbeitet   -> "Bitte Teilanforderungen ausfüllen"
      ein "zu klären"                -> zu klären
      alles entbehrlich              -> entbehrlich
      alles verwiesen                -> verwiesen
      nein + entbehrlich = alle      -> nein
      ja + entbehrlich + verwiesen   -> ja
      sonst                          -> teilweise
    """
    if bis < von:
        return STATUS_UNVOLLSTAENDIG
    r = f"${spalte}${von}:${spalte}${bis}"
    return (
        f'=IF(OR(COUNTBLANK({r})>0,COUNTIF({r},"unbearbeitet")=COUNTA({r})),'
        f'"{STATUS_UNVOLLSTAENDIG}",'
        f'IF(COUNTIF({r},"zu klären")>0,"zu klären",'
        f'IF(COUNTIF({r},"entbehrlich")=COUNTA({r}),"entbehrlich",'
        f'IF(COUNTIF({r},"verwiesen")=COUNTA({r}),"verwiesen",'
        f'IF(COUNTIF({r},"nein")+COUNTIF({r},"entbehrlich")=COUNTA({r}),"nein",'
        f'IF(COUNTIF({r},"ja")+COUNTIF({r},"entbehrlich")+COUNTIF({r},"verwiesen")'
        f'=COUNTA({r}),"ja","teilweise"))))))'
    )


def _blatt_uebersicht(wb: Workbook, reqs: list[Requirement]) -> None:
    """
    Auswertung je Praktik.

    Alle Zaehlungen filtern zusaetzlich auf Pruefpflichtig="ja". Sammelanforderungen
    fallen damit heraus - sie sind nur eine Klammer und wuerden Anzahl,
    Aufwandssumme und Erfuellungsgrad verfaelschen.
    """
    ws = wb.create_sheet(BLATT_UEBERSICHT, 1)

    header = [("Praktik", 10), ("Bezeichnung", 30),
              ("Anforderungen\n(pruefpflichtig)", 16), ("davon Teilanf.", 14),
              ("MUSS", 8), ("SOLLTE", 9), ("KANN", 8),
              ("normal-SdT", 12), ("erhoeht", 10), ("Aufwandssumme", 14)]
    header += [(f"Status: {s_}", 14) for s_ in UMSETZUNGSSTATUS]
    header += [("Erfuellungsgrad", 15)]
    _kopfzeile(ws, header)

    praktiken: dict[str, str] = {}
    for r in reqs:
        praktiken.setdefault(r.praktik_id, r.praktik_titel)

    n = len(reqs)
    q = f"'{BLATT_CHECK}'"
    C_PRAK = check_spalte("praktik_id")
    C_EBENE = check_spalte("ebene")
    C_PFLICHT = check_spalte("__pruefpflichtig")
    C_MODAL = check_spalte("modalverb")
    C_STUFE = check_spalte("schutzbedarfsstufe")
    C_AUFW = check_spalte("aufwand")
    C_STATUS = get_column_letter(len(CHECK_KERN) + 1)
    SP_STATUS_1 = 11
    SP_ERF = SP_STATUS_1 + len(UMSETZUNGSSTATUS)

    def rng(col: str) -> str:
        return f"{q}!${col}$2:${col}${n + 1}"

    for i, (pid, titel) in enumerate(sorted(praktiken.items()), start=2):
        ws.cell(row=i, column=1, value=pid).font = ZELL_FONT
        ws.cell(row=i, column=2, value=titel).font = ZELL_FONT
        basis = f'{rng(C_PRAK)},$A{i},{rng(C_PFLICHT)},"ja"'

        ws.cell(row=i, column=3, value=f"=COUNTIFS({basis})")
        ws.cell(row=i, column=4, value=f'=COUNTIFS({basis},{rng(C_EBENE)},">0")')
        for j, mv in enumerate(["MUSS", "SOLLTE", "KANN"], start=5):
            ws.cell(row=i, column=j, value=f'=COUNTIFS({basis},{rng(C_MODAL)},"{mv}")')
        for j, sl in enumerate(["normal-SdT", "erhöht"], start=8):
            ws.cell(row=i, column=j, value=f'=COUNTIFS({basis},{rng(C_STUFE)},"{sl}")')
        ws.cell(row=i, column=10, value=f"=SUMIFS({rng(C_AUFW)},{basis})")
        for j, st in enumerate(UMSETZUNGSSTATUS, start=SP_STATUS_1):
            ws.cell(row=i, column=j, value=f'=COUNTIFS({basis},{rng(C_STATUS)},"{st}")')

        erf = get_column_letter(SP_STATUS_1 + UMSETZUNGSSTATUS.index("ja"))
        entb = get_column_letter(SP_STATUS_1 + UMSETZUNGSSTATUS.index("entbehrlich"))
        ws.cell(row=i, column=SP_ERF, value=f"=IFERROR(({erf}{i}+{entb}{i})/C{i},0)")

        for j in range(3, SP_ERF + 1):
            c = ws.cell(row=i, column=j)
            c.font = ZELL_FONT
            c.border = RAHMEN
        ws.cell(row=i, column=SP_ERF).number_format = "0.0%"

    summe = len(praktiken) + 2
    ws.cell(row=summe, column=2, value="Gesamt").font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
    for j in range(3, SP_ERF):
        col = get_column_letter(j)
        c = ws.cell(row=summe, column=j, value=f"=SUM({col}2:{col}{summe - 1})")
        c.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)
    erf = get_column_letter(SP_STATUS_1 + UMSETZUNGSSTATUS.index("ja"))
    entb = get_column_letter(SP_STATUS_1 + UMSETZUNGSSTATUS.index("entbehrlich"))
    g = ws.cell(row=summe, column=SP_ERF,
                value=f"=IFERROR(({erf}{summe}+{entb}{summe})/C{summe},0)")
    g.number_format = "0.0%"
    g.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=True)

    hinweis = ws.cell(row=summe + 2, column=2,
                      value="Sammelanforderungen sind nicht gezaehlt - ihr Status wird "
                            "aus den Teilanforderungen aggregiert.")
    hinweis.font = Font(name=FONT, size=design.AKTIV.groesse_daten - 2, italic=True, color="808080")


def _blatt_legende(wb: Workbook) -> None:
    ws = wb.create_sheet(BLATT_LEGENDE)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96

    inhalte = [
        ("Bedienung", ""),
        ("Gelbe Zellen", "Nur diese Spalten werden ausgefuellt. Alles andere stammt aus dem "
                         "BSI-Katalog und wird beim naechsten Lauf ueberschrieben."),
        ("Blatt 01_Uebersicht", "Rechnet automatisch aus 03_Grundschutz-Check. Nicht manuell aendern."),
        ("Blatt 00_Metadaten", "Herkunftsnachweis. Bei Rueckfragen im Audit ist das der Beleg, "
                               "auf welchem Katalogstand die Bewertung beruht."),
        ("", ""),
        ("Modalverb", ""),
        ("MUSS", "Unbedingt zu erfuellen."),
        ("SOLLTE", "Zu erfuellen; Abweichung ist zu begruenden und zu dokumentieren."),
        ("KANN", "Optional, situationsabhaengig."),
        ("", ""),
        ("Schutzbedarfsstufe", ""),
        ("normal-SdT", "Stand der Technik bei normalem Schutzbedarf."),
        ("erhöht", "Zusaetzlich bei erhoehtem Schutzbedarf."),
        ("", ""),
        ("Schutzziele", "0 = kein Bezug, 1 = Beitrag, 2 = wesentlicher Beitrag"),
        ("Aufwand", "0 (kein zusaetzlicher Aufwand) bis 5 (sehr hoch)"),
        ("", ""),
        ("Umsetzungsstatus", " / ".join(UMSETZUNGSSTATUS)),
        ("entbehrlich", "Anforderung ist im konkreten Informationsverbund nicht anwendbar. "
                        "Begruendung ist Pflicht."),
    ]
    for i, (k, v) in enumerate(inhalte, start=1):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=not v)
        b.font = ZELL_FONT
        b.alignment = Alignment(wrap_text=True, vertical="top")


# ------------------------------------------------------------------- Einstiege


def erzeuge_mappe(snap: CatalogSnapshot, ziel: Path,
                  filter_stufe: str | None = None) -> Path:
    """GENERATE-Modus: vollstaendige Arbeitsmappe von Grund auf."""
    reqs = snap.requirements
    if filter_stufe:
        reqs = filter_mit_ahnen(reqs, lambda r: r.schutzbedarfsstufe == filter_stufe)
        pflicht = sum(1 for r in reqs if r.pruefpflichtig)
        log.info("Filter Schutzbedarfsstufe=%s -> %d Zeilen (%d pruefpflichtig, "
                 "%d Vorfahren als Kontext)", filter_stufe, len(reqs), pflicht,
                 len(reqs) - pflicht)

    wb = Workbook()
    wb.remove(wb.active)
    _blatt_anforderungen(wb, reqs)
    _blatt_check(wb, reqs)
    _blatt_uebersicht(wb, reqs)
    _blatt_legende(wb)
    _blatt_metadaten(wb, snap)
    wb._sheets.sort(key=lambda s: s.title)  # 00_ .. 04_

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)
    log.info("Geschrieben: %s (%d Anforderungen)", ziel, len(reqs))
    return ziel


def befuelle_template(snap: CatalogSnapshot, template: Path, ziel: Path,
                      blattname: str, kopfzeile: int = 1,
                      filter_stufe: str | None = None) -> Path:
    """
    TEMPLATE-Modus: bestehende Vorlage laden und nur die Datenzeilen schreiben.

    Das Spalten-Mapping laeuft ueber die Kopfzeilentexte der Vorlage. Spalten,
    deren Kopftext keinem Katalogfeld entspricht (z.B. eigene Bewertungsspalten),
    bleiben unangetastet.
    """
    reqs = snap.requirements
    if filter_stufe:
        reqs = filter_mit_ahnen(reqs, lambda r: r.schutzbedarfsstufe == filter_stufe)

    wb = load_workbook(template)
    if blattname not in wb.sheetnames:
        raise KeyError(f"Blatt '{blattname}' fehlt in der Vorlage. Vorhanden: {wb.sheetnames}")
    ws = wb[blattname]

    kopf_zu_attr = {t: a for t, a, _ in SPALTEN}
    mapping: dict[int, str] = {}
    unbekannt: list[str] = []
    for spalte in range(1, ws.max_column + 1):
        kopf = ws.cell(row=kopfzeile, column=spalte).value
        if not kopf:
            continue
        kopf = str(kopf).strip()
        if kopf in kopf_zu_attr:
            mapping[spalte] = kopf_zu_attr[kopf]
        else:
            unbekannt.append(kopf)

    if not mapping:
        raise KeyError(
            "Keine Spaltenkoepfe der Vorlage konnten zugeordnet werden. "
            f"Erwartet u.a.: {list(kopf_zu_attr)[:5]} - gefunden: {unbekannt[:8]}"
        )
    log.info("Template-Mapping: %d Spalten zugeordnet, %d bleiben unberuehrt (%s)",
             len(mapping), len(unbekannt), ", ".join(unbekannt[:6]))

    for zi, r in enumerate(reqs, start=kopfzeile + 1):
        for spalte, attr in mapping.items():
            ws.cell(row=zi, column=spalte, value=getattr(r, attr))

    if BLATT_META in wb.sheetnames:
        del wb[BLATT_META]
    _blatt_metadaten(wb, snap)

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)
    log.info("Template befuellt: %s (%d Zeilen)", ziel, len(reqs))
    return ziel
