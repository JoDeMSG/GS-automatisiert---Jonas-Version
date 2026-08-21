"""
Zentrales Design-Profil.

Alle Farb-, Schrift- und Layoutwerte an EINER Stelle. Ein Wechsel auf
msg-Hausfarben ist damit ein Austausch weniger Konstanten, kein Umbau.

Werte des Profils BSI stammen aus der Referenzdatei
'A_1_Vorlage_Strukturanalyse_1_1_0.xlsx' (BSI-Original, Vorlagenversion 1.1.0,
Kompendium-Edition 2023) und wurden dort ausgelesen, nicht geschaetzt:

    Schrift            Aptos, 11 pt (Datenblaetter), 12 pt (Deckblatt)
    Kopfzeile          Fuellung #56A3BC, fett, Zeilenhoehe 30
    Sekundaertext      #6F6F6F (Einstufungshinweis, Blattueberschriften)
    Zellrahmen         KEINE
    Gitternetzlinien   ausgeblendet
    Randspalten        links/rechts je 2,1 Breite als optischer Rand
    Datenbereich       keine Hintergrundfarbe, auch nicht bei Eingabefeldern
    Spaltenbreiten     grosszuegig (30,6 / 35,4 / 44)

Bewusste Abweichung vom Original: Eingabefelder und berechnete Zellen werden
dezent hinterlegt. Das Original kommt ohne aus, weil dort JEDE Zelle im
Datenbereich ein Eingabefeld ist. In unserem Grundschutz-Check stehen
Katalogdaten (nicht anfassen), Eingabefelder (ausfuellen) und Rollup-Formeln
(berechnet) nebeneinander - ohne Unterscheidung wuerde in die falschen Zellen
geschrieben. Die Toene sind deshalb sehr blass gewaehlt.
"""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.styles import Alignment, Font, PatternFill


@dataclass(frozen=True)
class Designprofil:
    name: str

    schrift: str
    groesse_daten: float
    groesse_deckblatt: float

    # Farben als RGB-Hex ohne '#'
    akzent: str            # Kopfzeilen
    akzent_text: str       # Text auf Akzentflaeche
    sekundaertext: str     # Hinweise, Blattueberschriften
    eingabe: str           # Eingabefelder
    berechnet: str         # Rollup-/Formelzellen
    trennlinie: str

    kopfzeilenhoehe: float = 30.0
    datenzeilenhoehe: float = 15.0
    randspaltenbreite: float = 2.1
    zellrahmen: bool = False
    gitternetz: bool = False

    # Ampelfarben je Umsetzungsstatus (Fuellung, Schriftfarbe).
    # Bewusst gedeckte Toene: die Statusspalte steht neben Fliesstext, grelle
    # Farben wuerden das Blatt unruhig machen. "unbearbeitet" bleibt bewusst
    # ohne Fuellung - der Normalzustand soll nicht auffallen.
    status_farben: dict = field(default_factory=lambda: {
        "ja":           ("D6E9D4", "1E5B2E"),   # gruen
        "nein":         ("F4CCCC", "8C1D18"),   # rot
        "teilweise":    ("FFF2CC", "7F6000"),   # gelb
        "entbehrlich":  ("E4E4E4", "6F6F6F"),   # grau
        "verwiesen":    ("D9E6F2", "1F4E79"),   # blau
        "zu klären":    ("FCE0C8", "9C4500"),   # orange
    })

    def kopf_fill(self) -> PatternFill:
        return PatternFill("solid", fgColor=self.akzent)

    def kopf_font(self) -> Font:
        return Font(name=self.schrift, size=self.groesse_daten, bold=True,
                    color=self.akzent_text)

    def zell_font(self, bold: bool = False, italic: bool = False,
                  color: str | None = None) -> Font:
        return Font(name=self.schrift, size=self.groesse_daten, bold=bold,
                    italic=italic, color=color)

    def hinweis_font(self) -> Font:
        return Font(name=self.schrift, size=self.groesse_daten, bold=True,
                    color=self.sekundaertext)

    def eingabe_fill(self) -> PatternFill:
        return PatternFill("solid", fgColor=self.eingabe)

    def berechnet_fill(self) -> PatternFill:
        return PatternFill("solid", fgColor=self.berechnet)


# --------------------------------------------------------------------- Profile

BSI = Designprofil(
    name="BSI",
    schrift="Aptos",
    groesse_daten=11.0,
    groesse_deckblatt=12.0,
    akzent="56A3BC",          # Petrolblau, exakt aus der Referenzdatei
    akzent_text="FFFFFF",
    sekundaertext="6F6F6F",   # exakt aus der Referenzdatei
    eingabe="FDF6E3",         # sehr blasses Warmgelb
    berechnet="F0F0F0",       # sehr blasses Grau
    trennlinie="D8DEE3",
)

# PLATZHALTER bis Luzians verbindliche Werte vorliegen.
# Der msg-Markenton ist ein kraeftiges Rot; die hier eingetragenen Werte sind
# eine begruendete Annaeherung, KEINE offiziellen CI-Werte. Vor Auslieferung
# an Kunden durch die echten Hexcodes ersetzen.
MSG = Designprofil(
    name="msg (Platzhalter)",
    schrift="Aptos",
    groesse_daten=11.0,
    groesse_deckblatt=12.0,
    akzent="9C1006",
    akzent_text="FFFFFF",
    sekundaertext="6F6F6F",
    eingabe="FDF6E3",
    berechnet="F0F0F0",
    trennlinie="E0D5D3",
)

PROFILE: dict[str, Designprofil] = {"bsi": BSI, "msg": MSG}

AKTIV: Designprofil = BSI


def setze_profil(name: str) -> Designprofil:
    """
    Setzt das aktive Profil UND frischt die abgeleiteten Stilkonstanten auf.

    Ohne das Auffrischen bliebe ein Profilwechsel wirkungslos: excel.py wertet
    KOPF_FILL, ZELL_FONT usw. beim Import aus, also bevor die Kommandozeile
    ueberhaupt gelesen wurde.
    """
    global AKTIV
    if name not in PROFILE:
        raise KeyError(f"Unbekanntes Designprofil '{name}'. Verfuegbar: {sorted(PROFILE)}")
    AKTIV = PROFILE[name]

    from openpyxl.styles import Border, Side

    from . import excel as _excel

    _excel.FONT = AKTIV.schrift
    _excel.KOPF_FILL = AKTIV.kopf_fill()
    _excel.KOPF_FONT = AKTIV.kopf_font()
    _excel.ZELL_FONT = AKTIV.zell_font()
    _excel.EINGABE_FILL = AKTIV.eingabe_fill()
    _excel.ROLLUP_FILL = AKTIV.berechnet_fill()
    duenn = Side(style="thin", color=AKTIV.trennlinie)
    _excel.DUENN = duenn
    _excel.RAHMEN = (Border(left=duenn, right=duenn, top=duenn, bottom=duenn)
                     if AKTIV.zellrahmen else Border())

    # schema2023 importiert die Konstanten by-value - dort ebenfalls nachziehen
    try:
        from . import schema2023 as _s2023
        for attr in ("FONT", "KOPF_FILL", "KOPF_FONT", "ZELL_FONT",
                     "EINGABE_FILL", "ROLLUP_FILL", "RAHMEN"):
            if hasattr(_s2023, attr):
                setattr(_s2023, attr, getattr(_excel, attr))
    except ImportError:
        pass

    return AKTIV
