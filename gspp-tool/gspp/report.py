"""Aenderungsbericht als Excel-Blatt und als Markdown (fuer Ticket/Mail)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import design
from .diff import Diffbericht, Gewicht
from .excel import FONT, KOPF_FILL, KOPF_FONT, ZELL_FONT, _kopfzeile

AMPEL = {
    Gewicht.KRITISCH: PatternFill("solid", fgColor="F8CBAD"),
    Gewicht.RELEVANT: PatternFill("solid", fgColor="FFE699"),
    Gewicht.REDAKTION: PatternFill("solid", fgColor="E2EFDA"),
}


def _kuerze(t: str, n: int = 500) -> str:
    return t if len(t) <= n else t[: n - 1] + "…"


def schreibe_excel(bericht: Diffbericht, ziel: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Aenderungen"

    header = [("Anforderungs-ID", 16), ("Alte ID", 16), ("Praktik", 10), ("Art", 12), ("Gewicht", 14),
              ("Titel", 40), ("Feld", 22), ("Alt", 60), ("Neu", 60)]
    _kopfzeile(ws, header)

    zeile = 2
    for a in bericht.aenderungen:
        if not a.felder:
            werte = [a.anforderung_id, a.alte_id or "", a.praktik_id, a.art, a.max_gewicht.value,
                     a.titel, "", "", ""]
            for i, v in enumerate(werte, start=1):
                c = ws.cell(row=zeile, column=i, value=v)
                c.font = ZELL_FONT
                c.alignment = Alignment(vertical="top", wrap_text=i in (6, 8, 9))
            ws.cell(row=zeile, column=5).fill = AMPEL[a.max_gewicht]
            zeile += 1
            continue
        for f in a.felder:
            werte = [a.anforderung_id, a.alte_id or "", a.praktik_id, a.art, f.gewicht.value, a.titel,
                     f.feld, _kuerze(f.alt), _kuerze(f.neu)]
            for i, v in enumerate(werte, start=1):
                c = ws.cell(row=zeile, column=i, value=v)
                c.font = ZELL_FONT
                c.alignment = Alignment(vertical="top", wrap_text=i in (6, 8, 9))
            ws.cell(row=zeile, column=5).fill = AMPEL[f.gewicht]
            zeile += 1

    kopf = wb.create_sheet("Zusammenfassung", 0)
    kopf.column_dimensions["A"].width = 30
    kopf.column_dimensions["B"].width = 60
    zeilen = [
        ("Aenderungsbericht Grundschutz++", ""),
        ("Von Katalog-Version", bericht.von_version),
        ("Nach Katalog-Version", bericht.nach_version),
        ("Von Commit", (bericht.von_commit or "n/a")[:12]),
        ("Nach Commit", (bericht.nach_commit or "n/a")[:12]),
        ("", ""),
        ("Neu", bericht.zaehle("neu")),
        ("Entfallen", bericht.zaehle("entfallen")),
        ("Umnummeriert", bericht.zaehle("umnummeriert")),
        ("Geaendert", bericht.zaehle("geaendert")),
        ("davon kritisch", len(bericht.kritisch)),
    ]
    for i, (k, v) in enumerate(zeilen, start=1):
        kopf.cell(row=i, column=1, value=k).font = Font(name=FONT, size=design.AKTIV.groesse_daten, bold=not v)
        kopf.cell(row=i, column=2, value=v).font = ZELL_FONT
    kopf["A1"].fill = KOPF_FILL
    kopf["A1"].font = KOPF_FONT

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)
    return ziel


def schreibe_markdown(bericht: Diffbericht, ziel: Path) -> Path:
    z = [
        "# Aenderungsbericht Grundschutz++",
        "",
        f"- Von: `{bericht.von_version}` (Commit `{(bericht.von_commit or 'n/a')[:12]}`)",
        f"- Nach: `{bericht.nach_version}` (Commit `{(bericht.nach_commit or 'n/a')[:12]}`)",
        f"- Bilanz: {bericht.zusammenfassung()}",
        "",
    ]
    for art, titel in [("neu", "Neue Anforderungen"), ("entfallen", "Entfallene Anforderungen")]:
        posten = [a for a in bericht.aenderungen if a.art == art]
        if posten:
            z += [f"## {titel} ({len(posten)})", ""]
            z += [f"- `{a.anforderung_id}` {a.titel}" for a in posten]
            z += [""]

    umzuege = [a for a in bericht.aenderungen if a.art == "umnummeriert"]
    if umzuege:
        unveraendert = [a for a in umzuege if not a.felder]
        veraendert = [a for a in umzuege if a.felder]
        z += [f"## Umnummeriert ({len(umzuege)})",
              "",
              "Gleicher oder nahezu gleicher Inhalt unter neuer ID - keine inhaltliche "
              "Neubewertung noetig, nur Referenzen/Verweise auf die alte ID pruefen.", ""]
        if unveraendert:
            z += [f"**Inhalt unveraendert ({len(unveraendert)}):**", ""]
            z += [f"- `{a.alte_id}` → `{a.anforderung_id}` {a.titel}" for a in unveraendert]
            z += [""]
        if veraendert:
            z += [f"**Umnummeriert UND inhaltlich veraendert ({len(veraendert)}) "
                  "- hier lohnt ein genauerer Blick:**", ""]
            for a in veraendert:
                felder = ", ".join(f.feld for f in a.felder)
                z += [f"- `{a.alte_id}` → `{a.anforderung_id}` {a.titel} — {felder}"]
            z += [""]

    krit = [a for a in bericht.aenderungen if a.art == "geaendert"
            and a.max_gewicht is Gewicht.KRITISCH]
    if krit:
        z += [f"## Inhaltlich geaendert - Neubewertung erforderlich ({len(krit)})", ""]
        for a in krit:
            z += [f"### `{a.anforderung_id}` {a.titel}"]
            for f in a.felder:
                if f.gewicht is Gewicht.KRITISCH:
                    z += [f"- **{f.feld}**", f"  - alt: {_kuerze(f.alt, 300)}",
                          f"  - neu: {_kuerze(f.neu, 300)}"]
            z += [""]

    rest = [a for a in bericht.aenderungen if a.art == "geaendert"
            and a.max_gewicht is not Gewicht.KRITISCH]
    if rest:
        z += [f"## Weitere Aenderungen ({len(rest)})", ""]
        for a in rest:
            felder = ", ".join(f.feld for f in a.felder)
            z += [f"- `{a.anforderung_id}` {a.titel} — {felder}"]

    ziel.parent.mkdir(parents=True, exist_ok=True)
    ziel.write_text("\n".join(z), encoding="utf-8")
    return ziel
